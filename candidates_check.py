import datetime
import json

import openpyxl
import os
import re


def reading_edit_xlsx_candidate_data():
    """ Чтение и сортировка данных по кандидатам """

    # Загружаем файл
    wb = openpyxl.load_workbook('Сводная таблица кандидатов.xlsx')
    # Получаем первый лист
    sheet = wb.active

    # Проверяем, существует ли файл
    if not os.path.exists('Отредактированная таблица кандидатов.xlsx'):
        new_wb = openpyxl.Workbook()
        # Создаем новый файл
        new_wb.save('Отредактированная таблица кандидатов.xlsx')

    # Открываем новый файл
    new_wb = openpyxl.load_workbook(filename='Отредактированная таблица кандидатов.xlsx', read_only=False)
    # new_wb = openpyxl.load_workbook(filename='Таблица кандидатов для загрузки.xlsx', read_only=False)
    new_sheet = new_wb.active

    new_sheet.delete_rows(1, new_sheet.max_row)  # Удаляем все строки

    count_num = 0  # счетчик строк
    phone_numbers = []  # список телефонов
    data_out = []  # выходные данные

    # Перебираем строки в исходном файле
    for row in sheet.iter_rows():
        count_num += 1

        # Получаем значения из столбцов
        value1 = row[0].value  # фио
        value2 = row[1].value  # дата звонка
        value3 = row[2].value  # телефон
        value4 = row[3].value  # возраст
        value5 = row[4].value  # регион
        value6 = row[5].value  # длина прицепа
        value7 = row[6].value  # стаж на фурах
        value8 = row[7].value  # источник информации
        value9 = row[8].value  # кто привел
        value10 = row[9].value  # решение СБ
        value11 = row[10].value  # колонный
        value12 = row[11].value  # комментарий колонного
        value13 = row[12].value  # медосмотр
        value14 = row[13].value  # комментарий ОК

# -------------------- проверка даты звонка -----------------
        # проверяем является ли запись в ячейке датой
        if isinstance(value2, datetime.datetime):
            value2 = value2
        else:
            value2 = None

# --------------------- проверка возраста ------------------------
        # проверяем является ли запись в ячейке датой
        # если да, то высчитываем возраст кандидата
        if isinstance(value4, datetime.datetime):
            if value2:
                value4 = value2.year - value4.year

            else:
                today = datetime.date.today()
                value4 = today.year - value4.year

        else:
            # Удаление всех символов, кроме цифр, в ячейке с возрастом
            value4 = re.sub(r'\D', '', str(value4))

            # преобразуем возраст в число
            if len(value4) > 1:
                value4 = int(value4)
            else:
                value4 = None

# --------------------- проверка региона ------------------------
        if value5:
            if len(value5) > 1:
                pass
            else:
                value5 = None

# --------------------- проверка длины прицепа ------------------
        # удаляем только буквы (латиница и кириллица)
        value6 = re.sub(r'[a-zA-Zа-яА-Я]', '', str(value6))
        value6 = re.sub(r'[.+,]', '+', value6)

# --------------------- проверка стажа на фурах ------------------
        # удаляем буквы (латиница и кириллица) и символы
        value7 = re.sub(r'[a-zA-Zа-яА-Я +,.()\t]', '', str(value7))
        value7 = str(value7)

        if value7:
            if len(value7) == 4:
                today = datetime.date.today()
                value7 = today.year - int(value7)
            try:
                if len(value7) == 0:
                    value7 = None
            except TypeError:
                pass

            try:
                if len(value7) <= 2:
                    try:
                        value7 = int(value7)
                    except ValueError:
                        value7 = None
            except TypeError:
                pass

        if isinstance(value7, str):
            value7 = None

# ------------------- проверка решения СБ ------------------------
        # удаляем пробелы и символы
        value10 = re.sub(r'[0-9 +,.()]', '', str(value10))
        value10 = value10.lower()

        if 'чистый' in value10 or 'соглас' in value10 or 'норм' in value10 or 'полож' in value10:
            value10 = 'Рекомендую'

        elif ('сограничениями' in value10 or 'несогласован' in value10 or 'нерекомендован' in value10
              or 'отказ' in value10):
            value10 = 'Не рекомендую'
        else:
            value10 = None

# ------------------- проверка медосмотра ------------------------
        # удаляем пробелы и символы
        value13 = re.sub(r'[0-9 +,.()]', '', str(value13))
        value13 = value13.lower()

        if ('полный' in value13 and 'сам' in value13) or ('выписан' in value13 and 'полный' in value13):
            value13 = 'Полный сам'

        elif ('первичный' in value13 and 'сам' in value13) or ('выписан' in value13 and 'первичный' in value13):
            value13 = 'Первичный сам'

        elif ('полный' in value13 and 'мы' in value13) or ('полный' in value13 and 'покупаем' in value13):
            value13 = 'Полный мы'

        elif ('первичный' in value13 and 'мы' in value13) or ('первич' in value13 and 'покупаем' in value13):
            value13 = 'Первичный мы'

        else:
            value13 = None

# --------------------- проверка комментария ОК ------------------------
        if value14:
            try:
                if len(value14) > 1:
                    value14 = str(value14)
                else:
                    value5 = None
            except TypeError:
                value14 = str(value14)

# --------------------- проверка номера телефона ----------------
        # Удаление всех символов, кроме цифр, в ячейке с номером телефона
        value3 = re.sub(r'\D', '', str(value3))

        # берем только номера больше 10 цифр
        if len(value3) >= 10:

            # если длина номера 10 цифр, то добавляем к нему 8
            if len(value3) == 10:
                value3 = '8' + value3

            # берем первые 11 цифр
            value3 = value3[:11]

            # если первая цифра в номере 7, то меняем ее на 8
            if value3[0] == '7':
                value3 = '8' + value3[1:]

            if value3 not in phone_numbers:  # если телефон еще не в списке
                phone_numbers.append(value3)  # добавляем телефон в общий список

                # Записываем отредактированные данные в новый лист
                new_sheet.append([value1, value2, value3, value4, value5, value6, value7, value8, value9, value10,
                                  value11, value12, value13, value14])

                # if count_num <= 400:

                data_out.append(
                    {
                        'candidate_name': value1,
                        'call_date': value2.strftime("%Y-%m-%d %H:%M:%S") if value2 else None,
                        'candidate_phone': value3,
                        'candidate_age': value4,
                        'region': value5,
                        'trailer_length': value6,
                        'candidate_experience': value7,
                        'information_source': value8,
                        'referer': value9,
                        'security': value10,
                        'column_chief': value11,
                        'column_chief_comment': value12,
                        'medical_checkup': value13,
                        'hr_comment': value14,
                    }
                )

    # Сохраняем изменения
    new_wb.save('Отредактированная таблица кандидатов.xlsx')
    # new_wb.save('Таблица кандидатов для загрузки.xlsx')

    file_out_json = os.path.abspath(f'./candidates_from_excel.json')

    print(len(data_out))

    with open(file_out_json, 'w', encoding='utf-8') as file:
        json.dump(data_out, file, sort_keys=False, indent=4, ensure_ascii=False)


reading_edit_xlsx_candidate_data()
